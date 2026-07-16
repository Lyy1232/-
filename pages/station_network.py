"""氢基产品需求信息网络 — 数据源：国内氢需求企业数据库 Excel · 双向同步 · 公路路线"""
import os, json
import streamlit as st
import folium
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import openpyxl
import math
from folium import Popup, Tooltip, Circle, PolyLine
from folium.plugins import Fullscreen, MarkerCluster
from streamlit_folium import st_folium
from utils.data_loader import load_sites
from utils.ui import render_module_header
from utils.road_router import road_route_from_bases, road_spider_routes
from config.constants import TECH_COLORS, TECH_ZH, ECONOMIC_RADIUS_KM, TRANSPORT_MODES, GUOHUA_BASES_COST

EXCEL_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data", "国内氢需求企业数据库.xlsx"))

RAW_COLS = [
    "序号","企业名称","所属集团","省份","城市","行业","子行业",
    "用氢场景","氢气形态需求","当前用氢量\n(吨H2/年)","绿氢替代潜力\n(吨H2/年)",
    "需求等级","需求确定性\n(1-5)","匹配集团基地","距最近基地\n预估距离(km)",
    "关键政策驱动","备注/关键项目","数据类型","需求坐标-经度","需求坐标-纬度",
    "坐标描述","需求点数量",
]
CLEAN_COLS = [c.replace('\n','') if isinstance(c,str) else c for c in RAW_COLS]

SHOW_COLS = ["序号","企业名称","所属集团","省份","城市","行业","子行业",
             "用氢场景","氢气形态需求","当前用氢量(吨H2/年)","绿氢替代潜力(吨H2/年)",
             "需求等级","需求确定性(1-5)","匹配集团基地","距最近基地预估距离(km)",
             "关键政策驱动","数据类型","需求坐标-经度","需求坐标-纬度","坐标描述","需求点数量"]

COL_LABELS = {
    "序号":"序号","企业名称":"企业名称","所属集团":"所属集团",
    "省份":"省份","城市":"城市","行业":"行业","子行业":"子行业",
    "用氢场景":"用氢场景","氢气形态需求":"氢气形态需求",
    "当前用氢量(吨H2/年)":"当前用氢量(t/y)","绿氢替代潜力(吨H2/年)":"绿氢替代潜力(t/y)",
    "需求等级":"需求等级","需求确定性(1-5)":"需求确定性",
    "匹配集团基地":"匹配集团基地","距最近基地预估距离(km)":"预估距离(km)",
    "关键政策驱动":"关键政策驱动","数据类型":"数据类型",
    "需求坐标-经度":"经度","需求坐标-纬度":"纬度",
    "坐标描述":"坐标描述","需求点数量":"需求点数量",
}

TYPE_COLORS = {"需求企业":"#2563eb","加氢站":"#10b981","政府/区域":"#f59e0b"}
TYPE_ICONS  = {"需求企业":"building","加氢站":"gas-pump","政府/区域":"flag"}
ICON_COLORS = {"需求企业":"blue","加氢站":"green","政府/区域":"orange"}
BASE_COORDS = {"赤城":(115.83,40.91),"宁东":(106.57,38.15),"沧州":(116.84,38.30),"如东":(121.18,32.33)}


# ═══════════════════════ EXCEL ═══════════════════════
@st.cache_data(show_spinner=False, ttl=10)
def _read_excel_cached(_path: str) -> pd.DataFrame:
    return _read_excel_raw()

def _read_excel_raw() -> pd.DataFrame:
    if not os.path.exists(EXCEL_PATH):
        return pd.DataFrame()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True); ws = wb.active
    headers = [ws.cell(row=3, column=c).value for c in range(1, len(RAW_COLS)+1)]
    headers = [h.strip() if isinstance(h,str) else (h or f"c{i}") for i,h in enumerate(headers)]
    data = []
    for r in range(4, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, len(headers)+1)]
        v0 = row[0]
        if v0 is None: continue
        if isinstance(v0, str) and (v0.strip().startswith("坐标") or v0.strip().startswith("排序")): continue
        data.append(row)
    wb.close()
    df = pd.DataFrame(data, columns=headers)
    df.columns = [c.replace('\n','') if isinstance(c,str) else c for c in df.columns]
    for c in ["当前用氢量(吨H2/年)","绿氢替代潜力(吨H2/年)","需求坐标-经度","需求坐标-纬度","需求点数量"]:
        if c in df.columns: df[c] = pd.to_numeric(df[c], errors='coerce')
    return df

def _write_excel(df: pd.DataFrame):
    wb = openpyxl.load_workbook(EXCEL_PATH); ws = wb.active
    for r in range(4, ws.max_row + 1):
        for c in range(1, len(RAW_COLS) + 1): ws.cell(row=r, column=c).value = None
    for ri, (_, row) in enumerate(df.iterrows()):
        for ci, raw_col in enumerate(RAW_COLS):
            clean_col = raw_col.replace('\n','') if isinstance(raw_col,str) else raw_col
            val = row.get(clean_col)
            try:
                if pd.isna(val): val = None
            except: pass
            ws.cell(row=4 + ri, column=ci + 1).value = val
    wb.save(EXCEL_PATH); wb.close()
    _read_excel_cached.clear()


# ═══════════════════════ 地图 ═══════════════════════
def _build_map(sites, demand_rows, route_polylines=None, highlight_point=None):
    """构建地图。可选参数：route_polylines=[[[lat,lon],...],...], highlight_point=(lat,lon,name)"""
    m = folium.Map(location=[37.5, 113.0], zoom_start=5,
                   tiles="CartoDB positron",
                   attr="CartoDB", control_scale=True)
    Fullscreen().add_to(m)

    # 基地 + 公路半径蜘蛛网
    spider_group = folium.FeatureGroup(name="公路经济半径")
    spider_colors = ["#ef4444","#f59e0b","#2563eb","#10b981","#8b5cf6"]
    for si, site in enumerate(sites):
        color = TECH_COLORS.get(site.get("tech",""), "#64748b")
        # 细虚线参考圈（直线200km）
        Circle([site["lat"],site["lon"]], radius=ECONOMIC_RADIUS_KM*1000,
               color="#94a3b8", fill=False, weight=0.8, opacity=0.25,
               dash_array="5,8").add_to(spider_group)
        # 基地标记
        folium.Marker([site["lat"],site["lon"]],
                      icon=folium.Icon(color="darkgreen", icon="industry", prefix="fa"),
                      popup=Popup(f"<b>{site['name']}</b><br>{TECH_ZH.get(site.get('tech',''),site.get('tech',''))}<br>¥{site.get('cost_avg','')}/kg", max_width=180)).add_to(m)
        folium.map.Marker([site["lat"]+0.04, site["lon"]],
                          icon=folium.DivIcon(html=f'<div style="font-size:9px;font-weight:700;color:#1e293b;background:rgba(255,255,255,0.9);padding:1px 4px;border-radius:2px">🏭 {site["name"]}</div>')).add_to(m)
        # 公路蜘蛛线（缓存读取）
        spider = road_spider_routes(
            {"name": site["name"], "lat": site["lat"], "lon": site["lon"]},
            [{"name": d["name"], "lat": d["lat"], "lon": d["lon"]} for d in demand_rows],
            ECONOMIC_RADIUS_KM,
        )
        for sp in spider:
            PolyLine(sp["polyline"], color=spider_colors[si % len(spider_colors)],
                     weight=2.2, opacity=0.55, dash_array="4,4").add_to(spider_group)
    spider_group.add_to(m)

    # 需求点
    cluster = MarkerCluster(name="需求点").add_to(m)
    for d in demand_rows:
        lat, lon = d.get("lat"), d.get("lon")
        if lat is None or lon is None: continue
        if not (-90<=lat<=90 and -180<=lon<=180): continue
        dt = d.get("type","需求企业")
        c = TYPE_COLORS.get(dt,"#94a3b8")
        popup = f"""<div style="font-family:-apple-system,sans-serif;min-width:200px">
          <h4 style="margin:0;font-size:12px">{d.get('name','')[:25]}</h4>
          <span style="background:{c}20;color:{c};padding:1px 6px;border-radius:3px;font-size:9px;font-weight:600">{dt}</span>
          <hr style="margin:4px 0"><table style="font-size:10px;line-height:1.5">
          <tr><td style="color:#64748b">省份</td><td><b>{d.get('prov','')} · {d.get('city','')}</b></td></tr>
          <tr><td style="color:#64748b">集团</td><td>{d.get('group','')}</td></tr>
          <tr><td style="color:#64748b">行业</td><td>{d.get('industry','')}</td></tr>
          <tr><td style="color:#64748b">用氢量</td><td><b>{d.get('demand','')} t/y</b></td></tr>
          <tr><td style="color:#64748b">需求等级</td><td>{d.get('grade','')}</td></tr>
          </table></div>"""
        icon_style = "star" if d.get("name") == highlight_point else ICON_COLORS.get(dt,"gray")
        ico = "star" if d.get("name") == highlight_point else TYPE_ICONS.get(dt,"info-sign")
        folium.Marker([lat,lon],
                      icon=folium.Icon(color="red" if d.get("name")==highlight_point else ICON_COLORS.get(dt,"gray"),
                                       icon=ico, prefix="fa"),
                      popup=Popup(popup, max_width=300),
                      tooltip=Tooltip(f"{dt} · {d.get('name','')[:20]}")).add_to(cluster)

    # 公路路线
    if route_polylines:
        route_group = folium.FeatureGroup(name="公路路线")
        colors = ["#ef4444","#f59e0b","#2563eb","#10b981"]
        for i, pl in enumerate(route_polylines):
            PolyLine(pl, color=colors[i%4], weight=3.5, opacity=0.85, popup=Popup(f"路线{i+1}", max_width=100)).add_to(route_group)
        route_group.add_to(m)

    # 高亮需求点
    if highlight_point:
        folium.Marker([highlight_point[0], highlight_point[1]],
                      icon=folium.Icon(color="red", icon="star", prefix="fa"),
                      popup=Popup(f"<b>📍 {highlight_point[2]}</b>", max_width=200)).add_to(m)

    # Legend
    lr = "".join(f'<tr><td><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:{c}"></span></td><td style="font-size:10px;color:#334155">{t}</td></tr>' for t,c in TYPE_COLORS.items())
    lr += '<tr><td><span style="display:inline-block;width:14px;height:2px;background:#94a3b8;border-radius:1px;border:0.5px dashed #94a3b8"></span></td><td style="font-size:10px;color:#334155">直线200km参考</td></tr>'
    lr += '<tr><td><span style="display:inline-block;width:14px;height:2px;background:repeating-linear-gradient(90deg,#888 0 4px,transparent 4px 8px)"></span></td><td style="font-size:10px;color:#334155">公路路线</td></tr>'
    m.get_root().html.add_child(folium.Element(f"""<div style="position:fixed;bottom:18px;right:18px;z-index:9999;background:rgba(255,255,255,0.95);padding:8px 13px;border-radius:8px;box-shadow:0 2px 12px rgba(0,0,0,0.08);line-height:1.6">
      <b style="font-size:11px;color:#1e293b">图例</b><table style="margin-top:3px">{lr}</table></div>"""))
    folium.LayerControl().add_to(m)
    return m


# ═══════════════════════ RENDER ═══════════════════════
def render():
    render_module_header("氢基产品需求信息网络", "国内氢需求企业数据库 · 双向同步Excel · 公路路线查询 · OSM底图", badge="LIVE")

    df = _read_excel_cached(EXCEL_PATH)
    if df.empty:
        st.error(f"⚠️ 未找到数据文件：{EXCEL_PATH}")
        return
    sites = load_sites()
    total = len(df)

    tc = df["数据类型"].value_counts()
    ent_c, sta_c, gov_c = tc.get("需求企业",0), tc.get("加氢站",0), tc.get("政府/区域",0)
    coord_ok = (df["需求坐标-经度"].notna() & df["需求坐标-纬度"].notna()).sum()
    total_h2 = df["当前用氢量(吨H2/年)"].sum() if "当前用氢量(吨H2/年)" in df.columns else 0

    st.markdown('<p style="font-size:10px;color:#94a3b8;text-transform:uppercase;letter-spacing:.8px;margin:0 0 8px;font-weight:600">数据总览</p>', unsafe_allow_html=True)
    for col,(ico,val,lab,sub) in zip(st.columns(5),[
        ("📋",str(total),"数据总量",f"需求企业 {ent_c} · 加氢站 {sta_c} · 政府 {gov_c}"),
        ("🏭",str(ent_c),"需求企业","化工/交通/电力/冶金等"),
        ("⛽",str(sta_c),"加氢站","已建成+在建+规划"),
        ("🗺️",str(coord_ok),"含GPS坐标",f"覆盖率 {coord_ok/max(total,1)*100:.0f}%"),
        ("⚡",f"{total_h2:,.0f}","当前用氢量 t/y","需求企业口径"),
    ]):
        with col:
            st.markdown(f"""<div class="metric-card"><div class="mc-accent-bar" style="background:#2563eb"></div><div class="mc-icon">{ico}</div><div class="mc-value">{val}</div><div class="mc-label">{lab}</div><div class="mc-sub">{sub}</div></div>""", unsafe_allow_html=True)

    st.markdown('<div style="margin:16px 0"></div>', unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["📋 数据管理","🗺️ 需求网络地图","📊 统计分析"])

    # ═══════════════════ TAB 1: 数据管理 ═══════════════════
    with tab1:
        st.info('💡 **双向同步**：修改后点击「保存到Excel」写入文件；在 Excel 中直接编辑后刷新页面即同步。')
        c1,c2,c3 = st.columns(3)
        with c1:
            tf = st.multiselect("数据类型筛选",["需求企业","加氢站","政府/区域"], default=["需求企业","加氢站","政府/区域"], key="dem_tf")
        with c2:
            provs = sorted(df["省份"].dropna().unique().tolist()) if "省份" in df.columns else []
            pf = st.multiselect("省份筛选", provs, default=[], key="dem_pf")
        with c3:
            kw = st.text_input("🔍 企业名称搜索","",placeholder="输入关键词...", key="dem_kw")
        disp = df.copy()
        if tf: disp = disp[disp["数据类型"].isin(tf)]
        if pf: disp = disp[disp["省份"].isin(pf)]
        if kw: disp = disp[disp["企业名称"].str.contains(kw, na=False)]
        st.caption(f"显示 {len(disp)} / {total} 条 | 双击编辑 · 底部可增删行")
        avail = [c for c in SHOW_COLS if c in disp.columns]
        edited = st.data_editor(
            disp[avail].rename(columns=COL_LABELS), num_rows="dynamic", use_container_width=True, height=500, key="dem_editor",
            column_config={
                "序号": st.column_config.NumberColumn(width="small"),
                "当前用氢量(t/y)": st.column_config.NumberColumn(format="%.0f"),
                "绿氢替代潜力(t/y)": st.column_config.NumberColumn(format="%.0f"),
                "需求确定性": st.column_config.NumberColumn(min_value=1,max_value=5),
                "经度": st.column_config.NumberColumn(format="%.4f"),
                "纬度": st.column_config.NumberColumn(format="%.4f"),
                "需求点数量": st.column_config.NumberColumn(width="small"),
            })
        c1,c2,c3 = st.columns([1,1,3])
        with c1:
            if st.button("💾 保存到Excel", type="primary", use_container_width=True, key="dem_save"):
                rev = {v:k for k,v in COL_LABELS.items()}
                saved = edited.rename(columns={c: rev.get(c,c) for c in edited.columns})
                for cc in saved.columns:
                    if cc in df.columns and len(saved) == len(disp):
                        df.loc[disp.index, cc] = saved[cc].values
                _write_excel(df)
                st.success(f"✅ 已保存到 Excel · {len(saved)} 行")
                st.rerun()
        with c2:
            if st.button("🔄 刷新", use_container_width=True, key="dem_refresh"):
                _read_excel_cached.clear(); st.rerun()
        with c3:
            st.caption(f"📁 `{os.path.basename(EXCEL_PATH)}`")

    # ═══════════════════ TAB 2: 地图 + 公路路线 ═══════════════════
    with tab2:
        mt = st.multiselect("地图显示类型",["需求企业","加氢站","政府/区域"],
                            default=["需求企业","加氢站"], key="dem_mt")
        rows = []
        for _, r in df.iterrows():
            lon_val = r.get("需求坐标-经度"); lat_val = r.get("需求坐标-纬度")
            try:
                lon_val = float(lon_val) if lon_val is not None and not (isinstance(lon_val,float) and pd.isna(lon_val)) else None
                lat_val = float(lat_val) if lat_val is not None and not (isinstance(lat_val,float) and pd.isna(lat_val)) else None
            except: lon_val = lat_val = None
            if lat_val is None or lon_val is None: continue
            dt = str(r.get("数据类型",""))
            if dt not in mt: continue
            rows.append(dict(name=str(r.get("企业名称","")),type=dt,lat=lat_val,lon=lon_val,
                             prov=str(r.get("省份","")),city=str(r.get("城市","")),
                             group=str(r.get("所属集团","")),industry=str(r.get("行业","")),
                             demand=str(r.get("当前用氢量(吨H2/年)","")),grade=str(r.get("需求等级",""))))

        # ── 路线查询 ──
        route_search = st.text_input("🔍 搜索需求点查询公路路线", "", placeholder="输入企业名称关键词 → 筛选后选择 → 自动显示公路路线及到站成本", key="route_search")
        filtered_rows = [r for r in rows if route_search in r["name"]] if route_search else []
        route_data = None; highlight_info = None; map_rows = rows  # 默认显示全部

        if filtered_rows:
            sel_name = st.selectbox("匹配的需求点", [r["name"] for r in filtered_rows], key="route_sel")
            target = next((r for r in filtered_rows if r["name"] == sel_name), None)
            if target:
                with st.spinner(f"正在计算公路路线到 {target['name']}..."):
                    base_list = [{"name":s["name"],"lon":s["lon"],"lat":s["lat"],"cost_avg":GUOHUA_BASES_COST.get(s["name"],27.0)} for s in sites]
                    route_data = road_route_from_bases(base_list, target["lon"], target["lat"])
                if route_data:
                    highlight_info = (target["lat"], target["lon"], target["name"])
                    st.success(f"✅ {target['name']} → {len(route_data)} 个基地公路路线已计算")
                    rt_rows = []
                    for rd in route_data:
                        tm = TRANSPORT_MODES["长管拖车30MPa"]
                        transport_fee = round(rd["road_km"] * tm["cost_per_100km"] / 100, 1)
                        landed = round(rd["base_cost"] + transport_fee, 1)
                        straight = round(math.sqrt((rd["base_lon"]-target["lon"])**2+(rd["base_lat"]-target["lat"])**2)*111, 0)
                        rt_rows.append({
                            "基地": rd["base_name"], "出厂成本(元/kg)": rd["base_cost"],
                            "公路距离(km)": rd["road_km"], "直线距离(km)": straight,
                            "差(km)": round(rd["road_km"]-straight, 1),
                            "运输费(元/kg)": transport_fee, "到站价(元/kg)": landed,
                            "耗时(min)": rd["duration_min"],
                        })
                    st.dataframe(pd.DataFrame(rt_rows), use_container_width=True, hide_index=True)
                    best = min(rt_rows, key=lambda x: x["到站价(元/kg)"])
                    st.info(f"💡 **最优路线**：{best['基地']} → {target['name']}，公路 {best['公路距离(km)']}km，运输费 {best['运输费(元/kg)']} 元/kg，到站价 **{best['到站价(元/kg)']} 元/kg**")
                else:
                    st.warning("⚠️ OSRM 计算失败，请重试")

        # ── 地图 ──
        st.caption(f"地图显示 {len(map_rows)} 个需求点" + (f" · 🚛 公路路线已叠加" if route_data else ""))
        m = _build_map(sites, map_rows,
                       route_polylines=[r["polyline"] for r in route_data] if route_data else None,
                       highlight_point=highlight_info)
        st_folium(m, width="100%", height=580, returned_objects=[])

    # ═══════════════════ TAB 3: 统计 ═══════════════════
    with tab3:
        c1,c2 = st.columns(2)
        with c1:
            st.subheader("行业分布")
            if "行业" in df.columns:
                ic = df["行业"].value_counts().head(10)
                fig = px.bar(x=ic.index,y=ic.values,color=ic.index, color_discrete_sequence=px.colors.qualitative.Pastel, labels={"x":"行业","y":"企业数"})
                fig.update_layout(height=320,showlegend=False,plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig,use_container_width=True)
        with c2:
            st.subheader("需求等级分布")
            if "需求等级" in df.columns:
                gc = df["需求等级"].value_counts().reindex(["A","B","C","D"])
                fig = px.bar(x=gc.index,y=gc.values,color=gc.index, color_discrete_map={"A":"#ef4444","B":"#f59e0b","C":"#2563eb","D":"#94a3b8"}, labels={"x":"需求等级","y":"企业数"})
                fig.update_layout(height=320,showlegend=False,plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig,use_container_width=True)
        st.subheader("省份分布")
        if "省份" in df.columns:
            pc = df["省份"].value_counts().head(15)
            fig = px.bar(x=pc.index,y=pc.values,color=pc.index, color_discrete_sequence=px.colors.qualitative.Set3, labels={"x":"省份","y":"企业/站数"})
            fig.update_layout(height=350,showlegend=False,plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig,use_container_width=True)
        c1,c2 = st.columns(2)
        with c1:
            st.subheader("数据类型分布")
            fig = px.pie(names=tc.index,values=tc.values,color=tc.index, color_discrete_map=TYPE_COLORS,hole=0.45,height=300)
            fig.update_traces(textinfo="label+value")
            fig.update_layout(showlegend=False,plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig,use_container_width=True)
        with c2:
            st.subheader("用氢量 Top 10")
            if "当前用氢量(吨H2/年)" in df.columns:
                t10 = df.nlargest(10, "当前用氢量(吨H2/年)")
                for _, r in t10.iterrows():
                    h2val = r.get("当前用氢量(吨H2/年)",0)
                    st.markdown(f"""<div style="display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid #f1f5f9">
                      <span style="font-size:12px;font-weight:600;color:#0f172a">{str(r.get('企业名称',''))[:18]}</span>
                      <span style="font-size:10px;color:#64748b">{r.get('省份','')} · {h2val:,.0f} t/y</span></div>""", unsafe_allow_html=True)

    st.caption("v0.5.0 · 需求信息网络 · 公路路线查询 · OSM底图 · OSRM引擎")
